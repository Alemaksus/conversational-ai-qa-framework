"""Tests for the Excel test case matrix loader."""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from caqf.data import CaseModel, MatrixSchemaError, load_test_cases


def test_load_test_cases_from_template():
    """Test loading test cases from the template Excel file."""
    xlsx_path = Path("templates/test-case-matrix.xlsx")
    
    if not xlsx_path.exists():
        pytest.skip(f"Template file not found: {xlsx_path}")
    
    test_cases = load_test_cases(str(xlsx_path))
    
    assert isinstance(test_cases, list)
    assert len(test_cases) >= 10, f"Expected at least 10 test cases, got {len(test_cases)}"
    
    # Verify first test case
    first_case = test_cases[0]
    assert isinstance(first_case, CaseModel)
    assert first_case.test_case_id, "First test case should have non-empty test_case_id"
    assert first_case.expected_result, "First test case should have non-empty expected_result"


def test_load_test_cases_structure():
    """Test that loaded test cases have correct structure."""
    xlsx_path = Path("templates/test-case-matrix.xlsx")
    
    if not xlsx_path.exists():
        pytest.skip(f"Template file not found: {xlsx_path}")
    
    test_cases = load_test_cases(str(xlsx_path))
    
    if not test_cases:
        pytest.skip("No test cases found in template file")
    
    # Verify all test cases have required fields
    for test_case in test_cases:
        assert test_case.test_case_id
        assert test_case.scenario_id is not None
        assert test_case.component is not None
        assert test_case.test_description is not None
        assert test_case.test_type is not None
        assert test_case.priority is not None
        assert test_case.prerequisites is not None
        assert test_case.test_steps is not None
        assert test_case.expected_result is not None
        # actual_result, status, notes are optional (can be None)


def test_matrix_schema_error_missing_column():
    """Test that MatrixSchemaError is raised when a required column is missing."""
    # Create a temporary workbook with missing "Priority" column
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Test Cases"
        
        # Add headers (missing "Priority")
        headers = [
            "Test Case ID",
            "Scenario ID",
            "Component",
            "Test Description",
            "Test Type",
            # "Priority" is missing
            "Prerequisites",
            "Test Steps",
            "Expected Result",
            "Actual Result",
            "Status",
            "Notes",
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Add a sample row
        sheet.cell(row=2, column=1, value="TC-001")
        sheet.cell(row=2, column=2, value="SCENARIO-001")
        sheet.cell(row=2, column=3, value="Component1")
        sheet.cell(row=2, column=4, value="Test description")
        sheet.cell(row=2, column=5, value="Functional")
        # Missing Priority
        sheet.cell(row=2, column=6, value="None")
        sheet.cell(row=2, column=7, value="Step 1")
        sheet.cell(row=2, column=8, value="Expected result")
        
        workbook.save(tmp_path)
        workbook.close()
        
        # Attempt to load should raise MatrixSchemaError
        with pytest.raises(MatrixSchemaError) as exc_info:
            load_test_cases(tmp_path)
        
        assert "Priority" in str(exc_info.value) or "Missing required columns" in str(exc_info.value)
    
    finally:
        # Clean up
        Path(tmp_path).unlink(missing_ok=True)


def test_load_test_cases_empty_file():
    """Test loading from an empty workbook (should raise error or return empty list)."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        workbook = openpyxl.Workbook()
        workbook.save(tmp_path)
        workbook.close()
        
        # Should raise MatrixSchemaError due to missing columns
        with pytest.raises(MatrixSchemaError):
            load_test_cases(tmp_path)
    
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def test_load_test_cases_nonexistent_file():
    """Test that FileNotFoundError is raised for nonexistent file."""
    with pytest.raises(FileNotFoundError):
        load_test_cases("nonexistent-file.xlsx")


def test_load_test_cases_without_optional_columns():
    """Test that optional columns (Actual Result, Status, Notes) can be missing."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Test Cases"
        
        # Add headers without optional columns
        headers = [
            "Test Case ID",
            "Scenario ID",
            "Component",
            "Test Description",
            "Test Type",
            "Priority",
            "Prerequisites",
            "Test Steps",
            "Expected Result",
            # "Actual Result", "Status", "Notes" are missing (optional)
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=header)
        
        # Add a sample row
        sheet.cell(row=2, column=1, value="TC-001")
        sheet.cell(row=2, column=2, value="SCENARIO-001")
        sheet.cell(row=2, column=3, value="Component1")
        sheet.cell(row=2, column=4, value="Test description")
        sheet.cell(row=2, column=5, value="Functional")
        sheet.cell(row=2, column=6, value="High")
        sheet.cell(row=2, column=7, value="None")
        sheet.cell(row=2, column=8, value="Step 1")
        sheet.cell(row=2, column=9, value="Expected result")
        
        workbook.save(tmp_path)
        workbook.close()
        
        # Should load successfully without optional columns
        test_cases = load_test_cases(tmp_path)
        assert len(test_cases) == 1
        assert test_cases[0].test_case_id == "TC-001"
        assert test_cases[0].actual_result is None
        assert test_cases[0].status is None
        assert test_cases[0].notes is None
    
    finally:
        Path(tmp_path).unlink(missing_ok=True)

