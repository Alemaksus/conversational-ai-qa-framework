"""Loader for Excel test case matrix files."""

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from caqf.data.models import TestCase

# Required column headers (order does not matter)
REQUIRED_COLUMNS = {
    "Test Case ID",
    "Scenario ID",
    "Component",
    "Test Description",
    "Test Type",
    "Priority",
    "Prerequisites",
    "Test Steps",
    "Expected Result",
    "Actual Result",
    "Status",
    "Notes",
}


class MatrixSchemaError(ValueError):
    """Raised when the Excel matrix does not contain required columns."""

    pass


def _get_sheet(workbook: Workbook) -> Worksheet:
    """Get the appropriate sheet from the workbook.
    
    Prefers "Test Cases" sheet if present, otherwise uses the first sheet.
    """
    if "Test Cases" in workbook.sheetnames:
        return workbook["Test Cases"]
    return workbook.active


def _build_header_mapping(sheet: Worksheet) -> dict[str, int]:
    """Build a mapping from header names to column indices.
    
    Reads row 1, strips whitespace from header names, and returns
    a dictionary mapping header_name -> column_index (1-based).
    """
    header_row = sheet[1]
    mapping = {}
    
    for cell in header_row:
        if cell.value is not None:
            header_name = str(cell.value).strip()
            if header_name:
                mapping[header_name] = cell.column
    
    return mapping


def _validate_required_columns(mapping: dict[str, int]) -> None:
    """Validate that all required columns exist in the header mapping.
    
    Raises MatrixSchemaError if any required columns are missing.
    """
    missing = REQUIRED_COLUMNS - set(mapping.keys())
    if missing:
        raise MatrixSchemaError(
            f"Missing required columns: {', '.join(sorted(missing))}"
        )


def _get_cell_value(sheet: Worksheet, row: int, col: int) -> str:
    """Get cell value as string, converting None to empty string."""
    cell = sheet.cell(row=row, column=col)
    if cell.value is None:
        return ""
    return str(cell.value)


def _is_row_empty(sheet: Worksheet, row: int, header_mapping: dict[str, int]) -> bool:
    """Check if a row is completely empty."""
    for col in header_mapping.values():
        cell = sheet.cell(row=row, column=col)
        if cell.value is not None and str(cell.value).strip():
            return False
    return True


def load_test_cases(xlsx_path: str) -> list[TestCase]:
    """Load test cases from an Excel file.
    
    Args:
        xlsx_path: Path to the Excel file
        
    Returns:
        List of TestCase objects
        
    Raises:
        MatrixSchemaError: If required columns are missing
        FileNotFoundError: If the file does not exist
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")
    
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = _get_sheet(workbook)
    
    # Build header mapping
    header_mapping = _build_header_mapping(sheet)
    
    # Validate required columns
    _validate_required_columns(header_mapping)
    
    # Extract column indices for required fields
    test_case_id_col = header_mapping["Test Case ID"]
    scenario_id_col = header_mapping["Scenario ID"]
    component_col = header_mapping["Component"]
    test_description_col = header_mapping["Test Description"]
    test_type_col = header_mapping["Test Type"]
    priority_col = header_mapping["Priority"]
    prerequisites_col = header_mapping["Prerequisites"]
    test_steps_col = header_mapping["Test Steps"]
    expected_result_col = header_mapping["Expected Result"]
    actual_result_col = header_mapping.get("Actual Result")
    status_col = header_mapping.get("Status")
    notes_col = header_mapping.get("Notes")
    
    test_cases = []
    
    # Iterate rows starting from row 2
    row = 2
    while True:
        # Stop if Test Case ID is empty
        test_case_id = _get_cell_value(sheet, row, test_case_id_col).strip()
        if not test_case_id:
            break
        
        # Skip completely empty rows
        if _is_row_empty(sheet, row, header_mapping):
            row += 1
            continue
        
        # Extract all fields
        scenario_id = _get_cell_value(sheet, row, scenario_id_col).strip()
        component = _get_cell_value(sheet, row, component_col).strip()
        test_description = _get_cell_value(sheet, row, test_description_col).strip()
        test_type = _get_cell_value(sheet, row, test_type_col).strip()
        priority = _get_cell_value(sheet, row, priority_col).strip()
        prerequisites = _get_cell_value(sheet, row, prerequisites_col).strip()
        test_steps = _get_cell_value(sheet, row, test_steps_col).strip()
        expected_result = _get_cell_value(sheet, row, expected_result_col).strip()
        
        # Optional fields (may be None)
        actual_result = None
        if actual_result_col:
            value = _get_cell_value(sheet, row, actual_result_col).strip()
            actual_result = value if value else None
        
        status = None
        if status_col:
            value = _get_cell_value(sheet, row, status_col).strip()
            status = value if value else None
        
        notes = None
        if notes_col:
            value = _get_cell_value(sheet, row, notes_col).strip()
            notes = value if value else None
        
        test_case = TestCase(
            test_case_id=test_case_id,
            scenario_id=scenario_id,
            component=component,
            test_description=test_description,
            test_type=test_type,
            priority=priority,
            prerequisites=prerequisites,
            test_steps=test_steps,
            expected_result=expected_result,
            actual_result=actual_result,
            status=status,
            notes=notes,
        )
        
        test_cases.append(test_case)
        row += 1
    
    workbook.close()
    return test_cases

